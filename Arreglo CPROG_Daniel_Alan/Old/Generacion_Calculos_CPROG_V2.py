
# coding: utf-8

# # Calculo CPROG V2

# Este programa genera un archivo en Excel a partir de la plantilla Cargos CPROG.xlsx en el cual se hace el calculo y se compara con la información de base de datos de CPROG para los OR de la 15 (para hacerlo se debe actualizar en la misma plantilla los OR que han entrado en el esquema
# 
# Realizado por: Carolina María Gómez
# Fecha: 02-03-2020
# 
# 1. Verificar en la plantilla Cargos_CPROG, en la hoja FechaIniVig que esten los OR para los cuales se va a hacer cálculos de cargos CPROG
# 2. Correr Código Calculo CPROG
# 3. Correr Código Verificación CPROG -- Actualmente Pendiente
# 4. Correr Código Validación histórica CPROG -- Actualmente Pendiente

# 1. Carga de base de datos y definición de funciones

# In[1]:


import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from copy import copy
import pandas as pd
import pyodbc 
from datetime import datetime, time
from openpyxl import load_workbook


#os.chdir("\\archivosxm\TransaccionesdelMercado\LAC\SDL - Plan Perdidas\08 Macros-Progrmas")

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
None_border = Border(left=Side(style=None), 
                     right=Side(style=None), 
                     top=Side(style=None), 
                     bottom=Side(style=None))
Empty_style = PatternFill(start_color='F0000000',
                   end_color='F0000000',
                   fill_type='solid')

def ConexionMIDCarperLAC():
    cnxn = pyodbc.connect("Driver={SQL Server Native Client 11.0};Server=comedxmv323,3052;Database=BDCargosPerdLAC;Trusted_Connection=yes;")
    return cnxn

def ConexionMID():
    cnxn = pyodbc.connect("Driver={SQL Server Native Client 11.0};Server=comedxmv144,3052;Database=BDMIDXM;Trusted_Connection=yes;")
    return cnxn

def Consulta_arg_1(cnxn, nombre_query, vable_1, arg_1):
    fd = open('Querys_CPROG\\'+nombre_query,'r').read()
    fd = fd.replace( vable_1 , arg_1 )
    DF_result = pd.read_sql(fd,cnxn)
    return DF_result

def Consulta_agr_3(cnxn, nombre_query, vable_1, arg_1,vable_2, arg_2,vable_3, arg_3):
    fd = open('Querys_CPROG\\'+nombre_query,'r').read()
    fd = fd.replace( vable_1 , arg_1 )
    fd = fd.replace( vable_2 , arg_2 )
    fd = fd.replace( vable_3 , arg_3 )
    DF_result = pd.read_sql(fd,cnxn)
    return DF_result

def Str_desde_array(Array):
    str_i = str()
    for a_i in Array:
        str_i = str_i + "'"+str(a_i)+"',"
    str_i = str_i[:len(str_i)-1]
    return str_i

def Consulta_IPP_aa(Fecha_IPP_a):
    cnxn = ConexionMID()
    # cnxn = ConexionMIDSDL()
    str_ippa = Str_desde_array(Fecha_IPP_a)
    DF_Ippa = Consulta_arg_1(cnxn, 'IPP_aa_V2.sql', ":Fechas", str_ippa)
    DF_Ippa['Fecha'] = DF_Ippa['Fecha'].dt.strftime('%Y-%m-%d')
    return DF_Ippa

def Consulta_Agentes(Agentes_a):
    cnxn = ConexionMID()
    # cnxn = ConexionMIDSDL()
    str_agentes = Str_desde_array(Agentes_a)
    DF_Agentes = Consulta_arg_1(cnxn, 'Ung_Agentes.sql', ":Agentes", str_agentes)
    return DF_Agentes

def Consulta_Ung_CarPerLAC(Ung_a):
    cnxn = ConexionMIDCarperLAC()
    # cnxn = ConexionMIDSDL()
    str_Ung = Str_desde_array(Ung_a)
    DF_Ung_CAP = Consulta_arg_1(cnxn, 'Ung_CAP.sql', ":Ungs", str_Ung)
    return DF_Ung_CAP

def Copiar_DF_excel(DF_imprimir, ws, r_1, c_1, border_sel):
    r = r_1
    for row in DF_imprimir.itertuples():
        cells = list(row)
        cells = cells[1:]
        c = c_1
        for cell in cells:
          # Actualizamos la celda de la planilla
            celldata = ws.cell(row = r, column = c)
            celldata.border = border_sel
            #print(str(c)+ " , " + str(r))
            celldata.value = cell
            c += 1
        r +=1

def Copiar_hoja_Excel(hoja_origen, Nombre_hoja_2, Nombre_wb):
    origen = wb[hoja_origen]
    ws2 = wb.copy_worksheet(origen)
    ws2.title= Nombre_hoja_2 
    ws2.sheet_view.showGridLines = False
    wb.save(Nombre_wb)
    
def Consulta_Ipp_CPROG(Fechas):
    cnxn = ConexionMID()
    DF_Ippa = Consulta_arg_1(cnxn, 'IPP_aa_V2.sql', ":Fechas", str(Fechas)[1:len(str(Fechas))-1])
    DF_Ippa['Fecha'] = DF_Ippa['Fecha'].dt.strftime('%Y-%m-%d')
    return DF_Ippa   

def Copiar_Valor_Variable_Excel(nombre_variable, wb,Valor):
    dests1 = (wb.defined_names[nombre_variable]).destinations
    cells = []
    for title, coord in dests1:
        cells.append(wb[title][coord])
    cells[len(cells)-1].value =  Valor
#     print(cells[len(cells)-1].value)

def Consulta_Insumos_AIMCP(result_3, Age):
    cnxn = ConexionMID()
    # cnxn = ConexionMIDSDL()
    NMA = result_3[result_3['Operador de Red']==Age]['NMA'].unique()[0]
    Ung_query= "'" + str(result_3[result_3['Operador de Red']==Age]['UngID'].unique()[0]) + "'"
    insumos_AIMCP = Consulta_agr_3(cnxn, 'Query_Insumos_AIMCP.sql', "@FI" , "'2019-05-01'","@NMA", str(NMA),"@agente", Ung_query)
    insumos_AIMCP['Fecha'] = pd.to_datetime(insumos_AIMCP['Fecha'])
    insumos_AIMCP['Fecha'] = insumos_AIMCP['Fecha'].dt.strftime('%Y-%m-%d')
    return insumos_AIMCP, Ung_query

def Consulta_Insumos_AIMCP_V2(result_3, Age):
    cnxn = ConexionMID()
    # cnxn = ConexionMIDSDL()
    NMA = result_3[result_3['Operador de Red']==Age]['NMA'].unique()[0]
    Ung_query= "'" + str(result_3[result_3['Operador de Red']==Age]['UngID'].unique()[0]) + "'"
    insumos_AIMCP = Consulta_agr_3(cnxn, 'Query_Insumos_AIMCP_V2.sql', "@FI" , "'2019-05-01'","@NMA", str(NMA),"@agente", Ung_query)
    insumos_AIMCP['Fecha'] = pd.to_datetime(insumos_AIMCP['Fecha'])
    insumos_AIMCP['Fecha'] = insumos_AIMCP['Fecha'].dt.strftime('%Y-%m-%d')
    return insumos_AIMCP, Ung_query
    
def Consulta_Insumos_CPROG(Mes, Ung_query):
    cnxn = ConexionMIDCarperLAC()
    insumos_CPROG = Consulta_agr_3(cnxn, 'Query_Insumos_CPROG.sql', "@FI" , str("'"+Mes+"'"),"@FF" , str("'"+Mes+"'"),"@agente", Ung_query)
    return insumos_CPROG

def Consulta_Insumos_CPROG_DGP(Mes, Ung_query):
    cnxn = ConexionMID()
    insumos_CPROG_DGP = Consulta_agr_3(cnxn, 'Query_Insumos_CPROG_DGP.sql', "@FI" , str("'"+Mes+"'"),"@FF" , str("'"+Mes+"'"),"@agente", Ung_query)
    return insumos_CPROG_DGP

def Consulta_Insumos_CPROG_DGP_V2(Mes, Ung_query):
    cnxn = ConexionMID()
    insumos_CPROG_DGP = Consulta_agr_3(cnxn, 'Query_Insumos_CPROG_DGP_V2.sql', "@FI" , str("'"+Mes+"'"),"@FF" , str("'"+Mes+"'"),"@agente", Ung_query)
    return insumos_CPROG_DGP

def Consulta_Calculo_CPROG(Ung_query):
    cnxn = ConexionMIDCarperLAC()
    Calculo_CPROG = Consulta_arg_1(cnxn, 'Query_Calculo_CPROG.sql', "@agente", Ung_query)
    return Calculo_CPROG

def Consulta_mapeo_nombre_agente(Ung_query):
    cnxn = ConexionMID()
    consulta_mapeo_agente = Consulta_arg_1(cnxn, 'Mapeo_agentes_nombre.sql', ":Agentes", Ung_query)
    return consulta_mapeo_agente

def Limpiar_variables(ws,r_i,r_f,c_i,c_f,border_sel = None_border,fill_style=Empty_style):
#     print("Limpiar Vbles",r_i,r_f,c_i,c_f)
    for r in range(r_i,r_f):
        for c in range(c_i,c_f):
            # Actualizamos la celda de la planilla
            celldata = ws.cell(row = r, column = c)
            celldata.border = border_sel
#             print(celldata)
            celldata.value = " "
            
def Limpiar_plantilla(ws):
    celldata = ws.cell(row = 3, column = 3)
    celldata.value = " "

    Limpiar_variables(ws,7,10,3,4,thin_border)
    Limpiar_variables(ws,12,13,2,5,thin_border)
    Limpiar_variables(ws,17,18,2,10,thin_border)
    Limpiar_variables(ws,24,44,2,8)
    Limpiar_variables(ws,23,24,2,7,thin_border)
    for i in range(7, 8):
        for j in range(24,44):
            copyStyle(ws.cell(row = j, column = i), ws.cell(row = 1, column = 1))


def copyStyle(newCell, cell):
    if cell.has_style:
        newCell.style = copy(cell.style)
        newCell.font = copy(cell.font)
        newCell.border = copy(cell.border)
        newCell.fill = copy(cell.fill)
        newCell.number_format = copy(cell.number_format)
        newCell.protection = copy(cell.protection)
        newCell.alignment = copy(cell.alignment)

def copyStyle_formula(ws, newCell, cell):
#     if cell.has_style:
        newCell.style = copy(cell.style)
        newCell.font = copy(cell.font)
        newCell.border = copy(cell.border)
        newCell.fill = copy(cell.fill)
        newCell.number_format = copy(cell.number_format)
        newCell.protection = copy(cell.protection)
        newCell.alignment = copy(cell.alignment)
#         print(str(cell.value), str(cell.coordinate),str(newCell.coordinate))
        ws[str(newCell.coordinate)] =openpyxl.formula.translate.Translator(cell.value,origin=str(cell.coordinate)).translate_formula(str(newCell.coordinate))
#         newCell.value = copy(cell.coordinate)
    
def style_rango_formula(ws, fila_ini, fila_fin, col_ini, col_fin):
    for i in range(col_ini, col_fin+1):
        for j in range(fila_ini+1,fila_fin):
            copyStyle_formula(ws, ws.cell(row = j, column = i), ws.cell(row = fila_ini, column = i))
#             print("Celda",j,i)
    
def style_rango(ws, fila_ini, fila_fin, col_ini, col_fin):
    for i in range(col_ini, col_fin):
        for j in range(fila_ini+1,fila_fin):
            copyStyle(ws.cell(row = j, column = i), ws.cell(row = fila_ini, column = i))
    


# 2. Colsulta información sobre inicio de vigencia
# 
# Se almacena en la hoja FechaIniVig y FechaIniVig2

# In[2]:


#Se carga información de la hoja 'FechaIniVIg', en la cual se debe poner manualmente los OR que van ingresando a plan de perdidas
pd_1 = pd.read_excel('Cargos CPROG.xlsx', sheet_name = 'FechaIniVig', nrows= 50, usecols = 'A:D') #Ojo Esta limitado a 50 OR
# pd_1['Fecha IPPaa'] = pd_1['Fecha IPPaa'].dt.strftime('%Y-%m-%d')
# pd_1['Fecha Ini Vig'] = pd_1['Fecha Ini Vig'].dt.strftime('%Y-%m-%d')
Fecha_IPP_a  = pd_1['Fecha IPPaa'].unique()
DF_Ippa = Consulta_IPP_aa(Fecha_IPP_a)
#Se consulta el IPP de los meses de inicio que se requiere para el cálculo del AIMCP 
result = pd.merge(pd_1, DF_Ippa, left_on='Fecha IPPaa', right_on='Fecha',how = 'left')
del result['Fecha']
Agentes_a= pd_1['Operador de Red'].unique()
#Se consulta la UNG de los operadores ya que se requiere para las consultas en CARPERLAC
DF_Agentes = Consulta_Agentes(Agentes_a)
result_2 = pd.merge(result, DF_Agentes, left_on='Operador de Red', right_on='IDENTAGENT',how = 'left')
del result_2['IDENTAGENT']
Ung_a= result_2['UngID'].unique()
#Se consulta el CAP por Unge
DF_Ung_CAP = Consulta_Ung_CarPerLAC(Ung_a)
result_3 = pd.merge(result_2, DF_Ung_CAP, left_on='UngID', right_on='unidadNegocio',how = 'left')
del result_3['unidadNegocio']
result_3
#Se carga el libro CPROG
wb = load_workbook('Cargos CPROG.xlsx')
ws = wb['FechaIniVig']
#Se copian los resultados de las consultas en la Hoja FechaIniVig y se guarda el libro
Copiar_DF_excel(result_3, ws, 2, 1, thin_border)
wb.save('Cargos CPROG.xlsx')
#Se carga el valor del IPP de abril de 2019 para la referencia que especifica la resolución para el cálculo del AIMCP
pd_1 = pd.read_excel('Cargos CPROG.xlsx', sheet_name = 'FechaIniVig', nrows= 1, usecols = 'H:I') #Ojo Esta limitado a 50 OR
pd_1['Fecha'] = pd_1['Fecha'].dt.strftime('%Y-%m-%d')
Fecha_IPP_a  = pd_1['Fecha'].unique()
DF_Ippa = Consulta_IPP_aa(Fecha_IPP_a)
#Se copia el valor de la variable en la Hoja del libro
Copiar_Valor_Variable_Excel('ipp_abr_19', wb, DF_Ippa['IPP'].unique()[0])
# Se copia la hoja para dejar los valores calculados
Copiar_hoja_Excel('FechaIniVig', 'FechaIniVig2', 'Cargos CPROG.xlsx')
wb = load_workbook('Cargos CPROG.xlsx')
Limpiar_variables(wb['FechaIniVig'],2,2+result_3.shape[0],5,5+result_3.shape[1]-4)
wb.save('Cargos CPROG.xlsx')


# In[3]:

#Mes - Es el mes que se va a liquidar
#Mes = '2020-04-01'
Mes = input("Ingrese el mes a ejecutar (yyyy-mm-dd)")
Mes_str = Mes
tipo =input("Ingrese el tipo de ejecución (Definitivo o Preliminar)")
#tipo = 'Definitivo'
Mes = datetime.strptime(Mes,'%Y-%m-%d')
#Se toman todos los OR que hacen parte de la resolución 015 (Deben estar en la Hoja 'FechaIniVig')
Agentes = result_3['Operador de Red']
# print(Agentes)
result_3_copy = result_3

#Se evalua los agentes que aplican segun su fecha de inicio en la hoja 'FechaIniVig'
for i in range(len(Agentes)):
    #Se toma la fecha de inicio de vigencia del agente 
    fecha_ini_age = str(result_3[result_3['Operador de Red']==Agentes[i]]['Fecha Ini Vig'].unique()[0])
    date_1=datetime(int(fecha_ini_age[0:4]),int(fecha_ini_age[5:7]),int(fecha_ini_age[8:10]))
    carry, new_month=divmod(date_1.month+1, 12)
    #Se suma un mes para evaluar la fecha de inicio de cálculo (si inicia mes m se inicia calculo mes siguiente)
    age_inicio_calculo=date_1.replace(year=date_1.year+carry, month=new_month)
    #Se evalua si para la fecha que se esta corriendo se debe calcular para ese agente
    if str(age_inicio_calculo) <= str(Mes):
        pass
#         print('Se incluye OR',Agentes[i])
    else:
#         print('Se borra OR',Agentes[i],i)
        result_3_copy=result_3_copy.drop(i,axis=0)
    
Agentes = result_3_copy['Operador de Red']
#print(Agentes)


# 3. Consulta de información CPROG

# In[4]:


for Age in Agentes:
    #Age = Agentes[0]
    print("Agente:" + str(Age))
    
    #Calculo de mes m-2 para evaluar consulta de IPP
    carry, new_month=divmod(Mes.month-2, 12)
    if new_month ==0:
        new_month =12
        carry=-1
    mes_m_2=Mes.replace(year=Mes.year+carry, month=new_month)
    
    #Fechas para cálculo
    Fecha_m_2 = datetime.strftime(mes_m_2,'%Y-%m-%d')
    Fehca_m_o = '2017-12-01'
    Fecha_ipp_a = result_3[result_3['Operador de Red']==Age]['Fecha IPPaa'].unique()[0]

    Fechas =[Fecha_m_2, Fehca_m_o, Fecha_ipp_a]
    #Consulta de IPPs para las fechas Fecha_m_2, Fehca_m_o, Fecha_ipp_a
    DF_Ippa=Consulta_Ipp_CPROG(Fechas)
    
    wb = load_workbook('Cargos CPROG.xlsx')
    ws = wb['CargoCPROG_1']

    Copiar_Valor_Variable_Excel('mes_cargo', wb,Mes_str)
    Copiar_Valor_Variable_Excel('ipp_o', wb,DF_Ippa[DF_Ippa['Fecha']==Fehca_m_o]['IPP'].unique()[0])
    Copiar_Valor_Variable_Excel('ipp_a', wb,DF_Ippa[DF_Ippa['Fecha']==Fecha_ipp_a]['IPP'].unique()[0])
    Copiar_Valor_Variable_Excel('ipp_m_2', wb,DF_Ippa[DF_Ippa['Fecha']==Fecha_m_2]['IPP'].unique()[0])
    Copiar_Valor_Variable_Excel('tipo', wb,tipo)   
    
    wb.save('Cargos CPROG.xlsx')
    if Age == 'EPID':
        insumos_AIMCP, Ung_query = Consulta_Insumos_AIMCP_V2(result_3, Age)
    else:
        insumos_AIMCP, Ung_query = Consulta_Insumos_AIMCP(result_3, Age)
    Copiar_DF_excel(insumos_AIMCP, ws, 23, 2, thin_border)
    style_rango(ws, 23, 23+insumos_AIMCP.shape[0], 2, 2+insumos_AIMCP.shape[1])
    style_rango_formula(ws, 23, 23+insumos_AIMCP.shape[0], 7, 7)
    
    insumos_CPROG = Consulta_Insumos_CPROG(Mes_str, Ung_query)
    insumos_CPROG_1 =pd.merge(insumos_CPROG,result_3, how='left', left_on = 'unidadNegocioOr', right_on = 'UngID')
    del insumos_CPROG_1['Operador de Red']
    del insumos_CPROG_1['Fecha Ini Vig']
    del insumos_CPROG_1['Fecha IPPaa']
    del insumos_CPROG_1['IPP']
    del insumos_CPROG_1['UngID']
    del insumos_CPROG_1['CAP']
    
    mapeo_agente = Consulta_mapeo_nombre_agente(Ung_query)
    insumos_CPROG_1['fechatrabajo'] = mapeo_agente['nombreCorto']
    insumos_CPROG_1['unidadNegocioOr'] = mapeo_agente['nombre']
    if Age == 'EPID':
        insumos_CPROG_DGP = Consulta_Insumos_CPROG_DGP_V2(Mes_str, Ung_query)
    else:
        insumos_CPROG_DGP = Consulta_Insumos_CPROG_DGP(Mes_str, Ung_query)
    
    insumos_CPROG_1['ventaUsuaConectadosStn']=insumos_CPROG_DGP['MgVenta_UsConectadosSTNCargoPerdidas']
    insumos_CPROG_1['ventaComNoIncumbente']=insumos_CPROG_DGP['MgVentaComercializadorNoIncumbenteCargoPerdidas']
    insumos_CPROG_1['ventaComIncumbente']=insumos_CPROG_DGP['MgVentaComercializadorIncumbenteCargoPerdidas']
    Copiar_DF_excel(insumos_CPROG_1, ws, 17, 2, thin_border)
    Calculo_CPROG = Consulta_Calculo_CPROG(Ung_query)
    Calculo_CPROG['unidadNegocio'] = mapeo_agente['nombreCorto']
    Copiar_DF_excel(Calculo_CPROG, ws, 12, 2, thin_border)
    Copiar_hoja_Excel('CargoCPROG_1', 'CargoCPROG-'+str(Age), 'Cargos CPROG.xlsx')
    
Limpiar_plantilla(ws)
ws.sheet_view.showGridLines = False
wb.save('Cargos CPROG'+str(Mes_str)+str(tipo)+'_V2_2.xlsx')
wb.save('Cargos CPROG.xlsx')
g_sheet=wb.sheetnames
for i in range(2,len(g_sheet)):    
    wb.remove(wb[g_sheet[i]])
wb.save('Cargos CPROG.xlsx')
