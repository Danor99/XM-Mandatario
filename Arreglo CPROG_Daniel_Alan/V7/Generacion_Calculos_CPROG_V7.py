
# coding: utf-8

#Creado por : Carolina María Gómez (50414)
#Fecha modif: 2020-04-03
# Calculo CPROG V4
# Modificaciones versión 4
# En la versiona 4 se realizan ajustes sobre versionamiento en las consultas ya que se realizan cambios en la base de datos de CARPERLAC
# Se agrega la creación de archivos en la carpeta de ejecuciones
# Se crea libreria independiente con funciones para facilitar entendimiento del código 

# Este programa genera un archivo en Excel a partir de la plantilla Cargos CPROG.xlsx en el cual se hace el calculo y se compara con la información de base de datos de CPROG para los OR de la 15 (para hacerlo se debe actualizar en la misma plantilla los OR que han entrado en el esquema
# 
# Realizado por: Carolina María Gómez
# Fecha: 02-03-2020
# 
# 1. Verificar en la plantilla Cargos_CPROG, en la hoja FechaIniVig que esten los OR para los cuales se va a hacer cálculos de cargos CPROG
# 2. Correr Código Calculo CPROG
# 3. Correr Código Verificación CPROG -- Actualmente Pendiente
# 4. Correr Código Validación histórica CPROG -- Actualmente Pendiente

# 1. Carga de base de datos y definición de funciones

# In[1]:


import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
#from copy import copy
import pandas as pd
#import pyodbc 
from datetime import datetime, time
from openpyxl import load_workbook
import Utilidades as uti
import os
from tkinter import *
from tkcalendar import*
import Verificacion_CPROG as veri

#os.chdir("\\archivosxm\TransaccionesdelMercado\LAC\SDL - Plan Perdidas\08 Macros-Progrmas")

global Mes
global tipo_user
global verif

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
None_border = Border(left=Side(style=None), 
                     right=Side(style=None), 
                     top=Side(style=None), 
                     bottom=Side(style=None))
Empty_style = PatternFill(start_color='F0000000',
                   end_color='F0000000',
                   fill_type='solid')

def asignar():
    global Mes
    global tipo_user
    global verif
    Mes=entry9.get()
    verif=Verificacion.get()
    if TipoPrelim.get() == 1 and TipoDefin.get() == 0:
        tipo_user = "Preliminar"
        raiz_guide.destroy()
    elif TipoPrelim.get() == 0 and TipoDefin.get() == 1:
        tipo_user = "Definitivo"
        raiz_guide.destroy()       
#    tipo_user=tipo_usr.get()

# 2. Colsulta información sobre inicio de vigencia
# 
# Se almacena en la hoja FechaIniVig y FechaIniVig2

# In[2]:

#Se carga información de la hoja 'FechaIniVIg', en la cual se debe poner manualmente los OR que van ingresando a plan de perdidas
pd_1 = pd.read_excel('Cargos CPROG_V2.xlsx', sheet_name = 'FechaIniVig', nrows= 50, usecols = 'A:D') #Ojo Esta limitado a 50 OR
pd_1 = pd_1.dropna()
#Se consultas los IPPaa es decir IPP inicio de aplicación 015 - para actualización cargo AIMCP
Fecha_IPP_a  = pd_1['Fecha IPPaa'].dropna(axis=0).unique()
DF_Ippa = uti.Consulta_IPP_aa(Fecha_IPP_a)

#Se consulta el IPP de los meses de inicio que se requiere para el cálculo del AIMCP, se agregan al DF tomado de la hoja para completar la información 
result = pd.merge(pd_1, DF_Ippa, left_on='Fecha IPPaa', right_on='Fecha',how = 'left')
del result['Fecha']
Agentes_a= pd_1['Operador de Red'].unique()

#Se consulta la UNG de los operadores ya que se requiere para las consultas en CARPERLAC
DF_Agentes = uti.Consulta_Agentes(Agentes_a)
result_2 = pd.merge(result, DF_Agentes, left_on='Operador de Red', right_on='IDENTAGENT',how = 'left')
del result_2['IDENTAGENT']
Ung_a= result_2['UngID'].unique()

#Se consulta el CAP por Unge
DF_Ung_CAP = uti.Consulta_Ung_CarPerLAC(Ung_a)
result_3 = pd.merge(result_2, DF_Ung_CAP, left_on='UngID', right_on='unidadNegocio',how = 'left')
del result_3['unidadNegocio']
#print(result_3)

#Se carga el libro CPROG que es la plantilla para realizar el archivo similar al de publicación
wb = load_workbook('Cargos CPROG_V2.xlsx')
ws = wb['FechaIniVig']

result_3['Fecha Ini Vig'] = pd.to_datetime(result_3['Fecha Ini Vig'], format='%Y-%m-%d')
result_3['Fecha IPPaa'] = pd.to_datetime(result_3['Fecha IPPaa'], format='%Y-%m-%d')
result_3['Fecha Ini Vig'] = result_3['Fecha Ini Vig'].dt.strftime('%Y-%m-%d')
result_3['Fecha IPPaa'] = result_3['Fecha IPPaa'].dt.strftime('%Y-%m-%d')
#Se copian los resultados de las consultas en la Hoja FechaIniVig y se guarda el libro
result_3 = result_3.drop_duplicates(subset=['Operador de Red'])
result_3 = result_3.reset_index(drop=True)
uti.Copiar_DF_excel(result_3, ws, 2, 1, thin_border)
wb.save('Cargos CPROG_V2.xlsx')

#Se carga el valor del IPP de abril de 2019 para la referencia que especifica la resolución para el cálculo del AIMCP
pd_1 = pd.read_excel('Cargos CPROG_V2.xlsx', sheet_name = 'FechaIniVig', nrows= 1, usecols = 'I:J') #Ojo Esta limitado a 50 OR
pd_1['Fecha'] = pd_1['Fecha'].dt.strftime('%Y-%m-%d')
Fecha_IPP_a  = pd_1['Fecha'].unique()
DF_Ippa = uti.Consulta_IPP_aa(Fecha_IPP_a)

#Se copia el valor de la variable en la Hoja del libro
uti.Copiar_Valor_Variable_Excel('ipp_abr_19', wb, DF_Ippa['IPP'].unique()[0])
# Se copia la hoja para dejar los valores calculados
uti.Copiar_hoja_Excel('FechaIniVig', 'FechaIniVig2', 'Cargos CPROG_V2.xlsx', wb)
wb = load_workbook('Cargos CPROG_V2.xlsx')
uti.Limpiar_variables(wb['FechaIniVig'],2,2+result_3.shape[0],5,5+result_3.shape[1]-4)
wb.save('Cargos CPROG_V2.xlsx')

# In[3]:

#Crear una ventana para los datos

#Crear el origen del Guide
raiz_guide=Tk() 
raiz_guide.title("Valores para calculo CPROG") #Titulo de la ventana
raiz_guide.iconbitmap("XM-LOGO-RGB-01.ico") #Poner icono
raiz_guide.geometry("350x280") #Definir el tamaño de la ventana
raiz_guide.resizable(0,0) #Bloquear para que el usuario no la pueda editar
main_title = Label(text="Ingrese los valores a calcular de CPROG",font=("Arial",12)) #Poner titulo dentro de la ventana
main_title.place(x=30,y=10) #Ubicación del texto de las variables de ingreso
#main_title.pack()

#Poner texto para las variables de ingreso
fecha_label=Label(text="Fecha de calculo (YYYY-MM-DD)",font=("Arial",10)) 
fecha_label.place(x=50,y=50) #Ubicación del texto de las variables de ingreso

#Definir las variables de ingreso
mes_usr = StringVar()
TipoPrelim = IntVar()
TipoDefin = IntVar()
Verificacion = IntVar()

#Definir los campos de los datos
#Texto1=Entry(raiz_guide, textvariable=mes_usr,width="40")
#Texto1.place(x=50,y=80)

entry9 = DateEntry(raiz_guide, width=10, background='black', day=1, foreground='white', borderwidth=2, date_pattern='y-mm-dd')
entry9.place(x=130,y=80)

C1=Checkbutton(raiz_guide,text="Preliminar",variable = TipoPrelim, onvalue=1, offvalue=0, height=2, width=7)
C1.place(x=130,y=110)
C2=Checkbutton(raiz_guide,text="Definitivo",variable = TipoDefin, onvalue=1, offvalue=0, height=2, width=7)
C2.place(x=130,y=140)
C3=Checkbutton(raiz_guide,text="Verificacion",variable = Verificacion, onvalue=1, offvalue=0, height=2, width=10)
C3.place(x=120,y=170)

#Boton
calcular_btn=Button(raiz_guide,text="Calcular CPROG",width="30",height="2", command=asignar)
calcular_btn.place(x=60,y=210)
raiz_guide.mainloop()

#Mes - Es el mes que se va a liquidar
#Mes = '2020-04-01'
Mes_str = Mes

#tipo =input("Ingrese el tipo de ejecución (Definitivo o Preliminar)")
tipo = tipo_user
Mes = datetime.strptime(Mes,'%Y-%m-%d')

if tipo == "Preliminar":
    directorio = "\\\\archivosxm\\TransaccionesdelMercado\\LAC\\SDL_Plan_Perdidas\\03_OR_con_plan_perdidas\\00_Ejecuciones\\Preliminar_"+Mes_str[:4]+"\\"
    directorio_archivos ="\\\\memweb01\\Admon_Mcdo\\Liquidacion_LAC_015\\PublicacionArchivos\\PERDIDAS\\Estimado\\"+Mes_str[:4]+"\\"+ Mes_str[5:7]+"\\"
else:
    directorio = "\\\\archivosxm\\TransaccionesdelMercado\\LAC\\SDL_Plan_Perdidas\\03_OR_con_plan_perdidas\\00_Ejecuciones\\Definitivo_"+Mes_str[:4]+"\\"
    directorio_archivos ="\\\\memweb01\\Admon_Mcdo\\Liquidacion_LAC_015\\PublicacionArchivos\\PERDIDAS\\Definitivo\\"+Mes_str[:4]+"\\"+ Mes_str[5:7]+"\\"
try:
    os.stat(directorio)
except:
    os.mkdir(directorio)

directorio = directorio + Mes_str[:7]+"\\"
try:
    os.stat(directorio)
except:
    os.mkdir(directorio)

    
    

result_3.dropna(subset=['NMA'], axis=0, inplace=True)
#Se toman todos los OR que hacen parte de la resolución 015 (Deben estar en la Hoja 'FechaIniVig')
Agentes = result_3['Operador de Red']
print(Agentes)
result_3_copy = result_3

#Se evalua los agentes que aplican segun su fecha de inicio en la hoja 'FechaIniVig'
for i in range(len(Agentes)):
    #Se toma la fecha de inicio de vigencia del agente 
    fecha_ini_age = str(result_3[result_3['Operador de Red']==Agentes[i]]['Fecha Ini Vig'].unique()[0])
    date_1=datetime(int(fecha_ini_age[0:4]),int(fecha_ini_age[5:7]),int(fecha_ini_age[8:10]))
    carry, new_month=divmod(date_1.month+1, 12)
    #Se suma un mes para evaluar la fecha de inicio de cálculo (si inicia mes m se inicia calculo mes siguiente)
    if(new_month == 0):
        new_month = 1
    
    age_inicio_calculo=date_1.replace(year=date_1.year+carry, month=new_month)
    #Se evalua si para la fecha que se esta corriendo se debe calcular para ese agente
    if str(age_inicio_calculo) <= str(Mes):
        pass
#         print('Se incluye OR',Agentes[i])
    else:
#         print('Se borra OR',Agentes[i],i)
        result_3_copy=result_3_copy.drop(i,axis=0)
#Se reasign a agentes con los que aplican para el mes    
Agentes = result_3_copy['Operador de Red']
print(Agentes)


# 3. Consulta de información CPROG

# In[4]:

for Age in Agentes:
    #Age = Agentes[0]
    print("Agente:" + str(Age))
    
    #Calculo de mes m-2 para evaluar consulta de IPP
    carry, new_month=divmod(Mes.month-2, 12)
    if new_month ==0:
        new_month =12
        carry=-1
    mes_m_2=Mes.replace(year=Mes.year+carry, month=new_month)
    
    #Fechas para cálculo
    Fecha_m_2 = datetime.strftime(mes_m_2,'%Y-%m-%d')
    Fehca_m_o = '2017-12-01'
    Fecha_ipp_a = result_3[result_3['Operador de Red']==Age]['Fecha IPPaa'].unique()[0]

    Fechas =[Fecha_m_2, Fehca_m_o, Fecha_ipp_a]
    
    #Consulta de IPPs para las fechas Fecha_m_2, Fehca_m_o, Fecha_ipp_a
    DF_Ippa=uti.Consulta_Ipp_CPROG(Fechas)
    
    #Se carga el libro de CPROG
    wb = load_workbook('Cargos CPROG_V2.xlsx')
    ws = wb['CargoCPROG_1']
    
    # Se ponen los valores de Mes, IPPs y Tipo en la hoja de CPROG
    uti.Copiar_Valor_Variable_Excel('mes_cargo', wb,Mes_str)
    uti.Copiar_Valor_Variable_Excel('ipp_o', wb,DF_Ippa[DF_Ippa['Fecha']==Fehca_m_o]['IPP'].unique()[0])
    uti.Copiar_Valor_Variable_Excel('ipp_a', wb,DF_Ippa[DF_Ippa['Fecha']==Fecha_ipp_a]['IPP'].unique()[0])
    uti.Copiar_Valor_Variable_Excel('ipp_m_2', wb,DF_Ippa[DF_Ippa['Fecha']==Fecha_m_2]['IPP'].unique()[0])
    uti.Copiar_Valor_Variable_Excel('tipo', wb,tipo_user)   
    # Guarda el libro
    wb.save('Cargos CPROG_V2.xlsx')
    
    
    # Para EPID se tiene el caso especial de sumar los valores de Enertolima y Celsia Tolima por lo cual tiene una consulta especial para su caso
    # Para el resto de agentes se hace consulta AIMCP que es información de DGP 
    Fecha_ini_RemunPGP = result_3[result_3['Operador de Red']==Age]['Fecha Ini Vig'].iloc[0]
    Fecha_ini_RemunPGP = datetime.strptime(Fecha_ini_RemunPGP,'%Y-%m-%d')
    Fecha_actual_RemunPGP = datetime.strptime(Mes_str,'%Y-%m-%d')
    
      
    if Age == 'EPID':
        insumos_AIMCP, Ung_query = uti.Consulta_Insumos_AIMCP_V2(result_3, Age)
    elif ((Age == 'CMMD') or (Age == 'CSSD')):
        #Se agrega esto para los OR que aplica la resolución CREG 010 de 2020
        insumos_AIMCP, Ung_query = uti.Consulta_Insumos_AIMCP_V3(result_3, Age)
    else:
        insumos_AIMCP, Ung_query = uti.Consulta_Insumos_AIMCP(result_3, Age)
    
    estado = uti.Consulta_Estado_plan_CPROG(Ung_query)
    uti.Copiar_DF_excel(estado, ws, 4, 8, thin_border)
    
    #Se copia en la hoja de Excel los insumos del AIMCP 
    if ((Fecha_actual_RemunPGP- Fecha_ini_RemunPGP).days < 370):
        insumos_AIMCP['Fecha']= pd.to_datetime(insumos_AIMCP['Fecha'], format='%Y-%m-%d')
        #insumos_AIMCP['Fecha'] = insumos_AIMCP['Fecha'].dt.strftime('%Y-%m-%d')
        uti.Copiar_DF_excel(insumos_AIMCP, ws, 23, 2, thin_border)
        #Se organiza el estilo de los campos que solo tienen datos
        uti.style_rango(ws, 23, 23+insumos_AIMCP.shape[0], 2, 2+insumos_AIMCP.shape[1])
        #Se organiza y copia los estilos y las formulas de los insumos del AIMCP
        uti.style_rango_formula(ws, 23, 23+insumos_AIMCP.shape[0], 7, 7)
    
    #Se define la variable tipo para en la consulta en carperlac indicar que versión traer (0 Prelim o 1 Defin)
    if tipo_user == "Preliminar":
        tipo = "0"
    else:
        tipo = "1"
        
    #Consulta de los insumos de CARPERLAC (Cprog, NMA, AIMCP )
    insumos_CPROG = uti.Consulta_Insumos_CPROG(Mes_str, Ung_query, tipo)
    insumos_CPROG_1 =pd.merge(insumos_CPROG,result_3, how='left', left_on = 'unidadNegocioOr', right_on = 'UngID')
    
    del insumos_CPROG_1['Operador de Red']
    del insumos_CPROG_1['Fecha Ini Vig']
    del insumos_CPROG_1['Fecha IPPaa']
    del insumos_CPROG_1['IPP']
    del insumos_CPROG_1['UngID']
    del insumos_CPROG_1['CAP']
    
    #Consulta para poner el nombre del agente
    mapeo_agente = uti.Consulta_mapeo_nombre_agente_AJ(Ung_query)
    insumos_CPROG_1['fechatrabajo'] = mapeo_agente['nombreCorto']
    insumos_CPROG_1['unidadNegocioOr'] = mapeo_agente['nombre']
    
    #Caso especial para EPID
    #Consulta insumos DGP para mes actual
    if Age == 'EPID':
        insumos_CPROG_DGP = uti.Consulta_Insumos_CPROG_DGP_V2(Mes_str, Ung_query)
    else:
        insumos_CPROG_DGP = uti.Consulta_Insumos_CPROG_DGP(Mes_str, Ung_query)
    
    #Se reasignan los valores con los insumos de DGP
    insumos_CPROG_1['ventaUsuaConectadosStn']=insumos_CPROG_DGP['MgVenta_UsConectadosSTNCargoPerdidas']
    insumos_CPROG_1['ventaComNoIncumbente']=insumos_CPROG_DGP['MgVentaComercializadorNoIncumbenteCargoPerdidas']
    insumos_CPROG_1['ventaComIncumbente']=insumos_CPROG_DGP['MgVentaComercializadorIncumbenteCargoPerdidas']
    #Se copia en excel los DF
    uti.Copiar_DF_excel(insumos_CPROG_1, ws, 17, 2, thin_border)
    #Consulta de valores para CPROG
    Calculo_CPROG = uti.Consulta_Calculo_CPROG(Ung_query)
    Calculo_CPROG['unidadNegocio'] = mapeo_agente['nombreCorto']
    uti.Copiar_DF_excel(Calculo_CPROG, ws, 12, 2, thin_border)
    uti.Copiar_hoja_Excel('CargoCPROG_1', 'CargoCPROG-'+str(Age), 'Cargos CPROG_V2.xlsx',wb)
    
uti.Limpiar_plantilla(ws)
ws.sheet_view.showGridLines = False
wb.save(directorio+'Cargos CPROG_'+str(Mes_str)+"_"+str(tipo_user)+'_V6.xlsx')
wb.save('Cargos CPROG_V2.xlsx')
g_sheet=wb.sheetnames
for i in range(2,len(g_sheet)):    
    wb.remove(wb[g_sheet[i]])
wb.save('Cargos CPROG_V2.xlsx')


mapeo_meses = {'01': 'Jan','02': 'Feb','03': 'Mar','04': 'Apr','05': 'May','06': 'Jun','07': 'Jul','08': 'Aug','09': 'Sep','10': 'Oct','11': 'Nov','12': 'Dec',}

if verif:
    print("Realizando verificación de archivos generados por CARPERLAC")
    veri.verificacion_CPROG(Mes_str,tipo_user, mapeo_meses, DF_Agentes, directorio, directorio_archivos)
    print("Verificación archivos finalizada")
