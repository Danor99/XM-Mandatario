# -*- coding: utf-8 -*-
"""
Created on Fri Apr  3 09:23:08 2020

@author: 50414
"""
import pandas as pd
import openpyxl
import pyodbc 
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Alignment
from copy import copy

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
None_border = Border(left=Side(style=None), 
                     right=Side(style=None), 
                     top=Side(style=None), 
                     bottom=Side(style=None))
Empty_style = PatternFill(start_color='F0000000',
                   end_color='F0000000',
                   fill_type='solid')


def ConexionMIDCarperLAC():
    cnxn = pyodbc.connect("Driver={SQL Server Native Client 11.0};Server=comedxmv323,3052;Database=BDCargosPerdLAC;Trusted_Connection=yes;")
    return cnxn

def ConexionMID():
    cnxn = pyodbc.connect("Driver={SQL Server Native Client 11.0};Server=comedxmv144,3052;Database=BDMIDXM;Trusted_Connection=yes;")
    return cnxn

def Consulta_arg_1(cnxn, nombre_query, vable_1, arg_1):
    fd = open('Querys_CPROG\\'+nombre_query,'r').read()
    fd = fd.replace( vable_1 , arg_1 )
    DF_result = pd.read_sql(fd,cnxn)
    return DF_result

def Consulta_agr_3(cnxn, nombre_query, vable_1, arg_1,vable_2, arg_2,vable_3, arg_3):
    fd = open('Querys_CPROG\\'+nombre_query,'r').read()
    fd = fd.replace( vable_1 , arg_1 )
    fd = fd.replace( vable_2 , arg_2 )
    fd = fd.replace( vable_3 , arg_3 )
    DF_result = pd.read_sql(fd,cnxn)
    return DF_result

def Str_desde_array(Array):
    str_i = str()
    for a_i in Array:
        str_i = str_i + "'"+str(a_i)+"',"
    str_i = str_i[:len(str_i)-1]
    return str_i

def Consulta_IPP_aa(Fecha_IPP_a):
    cnxn = ConexionMID()
    # cnxn = ConexionMIDSDL()
    str_ippa = Str_desde_array(Fecha_IPP_a)
    DF_Ippa = Consulta_arg_1(cnxn, 'IPP_aa_V2.sql', ":Fechas", str_ippa)
    DF_Ippa['Fecha'] = DF_Ippa['Fecha'].dt.strftime('%Y-%m-%d')
    return DF_Ippa

def Consulta_Agentes(Agentes_a):
    cnxn = ConexionMID()
    # cnxn = ConexionMIDSDL()
    str_agentes = Str_desde_array(Agentes_a)
    DF_Agentes = Consulta_arg_1(cnxn, 'Ung_Agentes.sql', ":Agentes", str_agentes)
    return DF_Agentes

def Consulta_Ung_CarPerLAC(Ung_a):
    cnxn = ConexionMIDCarperLAC()
    # cnxn = ConexionMIDSDL()
    str_Ung = Str_desde_array(Ung_a)
    DF_Ung_CAP = Consulta_arg_1(cnxn, 'Ung_CAP.sql', ":Ungs", str_Ung)
    return DF_Ung_CAP

def Copiar_DF_excel(DF_imprimir, ws, r_1, c_1, border_sel):
    r = r_1
    for row in DF_imprimir.itertuples():
        cells = list(row)
        cells = cells[1:]
        c = c_1
        for cell in cells:
          # Actualizamos la celda de la planilla
            celldata = ws.cell(row = r, column = c)
            celldata.border = border_sel
            celldata.alignment = Alignment(wrapText=True)
            #print(str(c)+ " , " + str(r))
            celldata.value = cell
            c += 1
        r +=1

def Copiar_hoja_Excel(hoja_origen, Nombre_hoja_2, Nombre_wb, wb):
    origen = wb[hoja_origen]
    ws2 = wb.copy_worksheet(origen)
    ws2.title= Nombre_hoja_2 
    ws2.sheet_view.showGridLines = False
    wb.save(Nombre_wb)
    
def Consulta_Ipp_CPROG(Fechas):
    cnxn = ConexionMID()
    DF_Ippa = Consulta_arg_1(cnxn, 'IPP_aa_V2.sql', ":Fechas", str(Fechas)[1:len(str(Fechas))-1])
    DF_Ippa['Fecha'] = DF_Ippa['Fecha'].dt.strftime('%Y-%m-%d')
    return DF_Ippa   

def Copiar_Valor_Variable_Excel(nombre_variable, wb,Valor):
    dests1 = (wb.defined_names[nombre_variable]).destinations
    cells = []
    for title, coord in dests1:
        cells.append(wb[title][coord])
    cells[len(cells)-1].value =  Valor
#     print(cells[len(cells)-1].value)

def Consulta_Insumos_AIMCP(result_3, Age):
    cnxn = ConexionMID()
    # cnxn = ConexionMIDSDL()
    NMA = result_3[result_3['Operador de Red']==Age]['NMA'].unique()[0]
    Ung_query= "'" + str(result_3[result_3['Operador de Red']==Age]['UngID'].unique()[0]) + "'"
    insumos_AIMCP = Consulta_agr_3(cnxn, 'Query_Insumos_AIMCP.sql', "@FI" , "'2019-05-01'","@NMA", str(NMA),"@agente", Ung_query)
    insumos_AIMCP['Fecha'] = pd.to_datetime(insumos_AIMCP['Fecha'])
    insumos_AIMCP['Fecha'] = insumos_AIMCP['Fecha'].dt.strftime('%Y-%m-%d')
    return insumos_AIMCP, Ung_query

def Consulta_Insumos_AIMCP_V2(result_3, Age):
    cnxn = ConexionMID()
    # cnxn = ConexionMIDSDL()
    NMA = result_3[result_3['Operador de Red']==Age]['NMA'].unique()[0]
    Ung_query= "'" + str(result_3[result_3['Operador de Red']==Age]['UngID'].unique()[0]) + "'"
    insumos_AIMCP = Consulta_agr_3(cnxn, 'Query_Insumos_AIMCP_V2.sql', "@FI" , "'2019-05-01'","@NMA", str(NMA),"@agente", Ung_query)
    insumos_AIMCP['Fecha'] = pd.to_datetime(insumos_AIMCP['Fecha'])
    insumos_AIMCP['Fecha'] = insumos_AIMCP['Fecha'].dt.strftime('%Y-%m-%d')
    return insumos_AIMCP, Ung_query

def Consulta_Insumos_AIMCP_V3(result_3, Age):
    cnxn = ConexionMID()
    # cnxn = ConexionMIDSDL()
    NMA = result_3[result_3['Operador de Red']==Age]['NMA'].unique()[0]
    Ung_query= "'" + str(result_3[result_3['Operador de Red']==Age]['UngID'].unique()[0]) + "'"
    insumos_AIMCP = Consulta_agr_3(cnxn, 'Query_Insumos_AIMCP.sql', "@FI" , "'2021-05-01'","@NMA", str(NMA),"@agente", Ung_query)
    insumos_AIMCP['Fecha'] = pd.to_datetime(insumos_AIMCP['Fecha'])
    insumos_AIMCP['Fecha'] = insumos_AIMCP['Fecha'].dt.strftime('%Y-%m-%d')
    return insumos_AIMCP, Ung_query
    
def Consulta_Insumos_CPROG(Mes, Ung_query, tipo=1):
    cnxn = ConexionMIDCarperLAC()
    insumos_CPROG = Consulta_agr_3(cnxn, 'Query_Insumos_CPROG.sql', "@FI" , str("'"+Mes+"'"),"@tipo" , tipo,"@agente", Ung_query)
    return insumos_CPROG

def Consulta_Estado_plan_CPROG(Ung_query):
    cnxn = ConexionMIDCarperLAC()
    insumos_CPROG = Consulta_arg_1(cnxn, 'Estado_plan.sql',"@agente", Ung_query)
    return insumos_CPROG

def Consulta_Insumos_CPROG_DGP(Mes, Ung_query):
    cnxn = ConexionMID()
    insumos_CPROG_DGP = Consulta_agr_3(cnxn, 'Query_Insumos_CPROG_DGP.sql', "@FI" , str("'"+Mes+"'"),"@FF" , str("'"+Mes+"'"),"@agente", Ung_query)
    return insumos_CPROG_DGP

def Consulta_Insumos_CPROG_DGP_V2(Mes, Ung_query):
    cnxn = ConexionMID()
    insumos_CPROG_DGP = Consulta_agr_3(cnxn, 'Query_Insumos_CPROG_DGP_V2.sql', "@FI" , str("'"+Mes+"'"),"@FF" , str("'"+Mes+"'"),"@agente", Ung_query)
    return insumos_CPROG_DGP

def Consulta_Calculo_CPROG(Ung_query):
    cnxn = ConexionMIDCarperLAC()
    Calculo_CPROG = Consulta_arg_1(cnxn, 'Query_Calculo_CPROG.sql', "@agente", Ung_query)
    return Calculo_CPROG

def Consulta_mapeo_nombre_agente(Ung_query):
    cnxn = ConexionMID()
    consulta_mapeo_agente = Consulta_arg_1(cnxn, 'Mapeo_agentes_nombre.sql', ":Agentes", Ung_query)
    return consulta_mapeo_agente

def Consulta_mapeo_nombre_agente_AJ(Ung_query):
    cnxn = ConexionMID()
    Consulta_mapeo_nombre_agente_AJ = Consulta_arg_1(cnxn, 'Mapeo_agentes_nombre_AJ.sql', ":Agentes", Ung_query)
    return Consulta_mapeo_nombre_agente_AJ

def Limpiar_variables(ws,r_i,r_f,c_i,c_f,border_sel = None_border,fill_style=Empty_style):
#     print("Limpiar Vbles",r_i,r_f,c_i,c_f)
    for r in range(r_i,r_f):
        for c in range(c_i,c_f):
            # Actualizamos la celda de la planilla
            celldata = ws.cell(row = r, column = c)
            celldata.border = border_sel
#             print(celldata)
            celldata.value = " "
            
def Limpiar_plantilla(ws):
    celldata = ws.cell(row = 3, column = 3)
    celldata.value = " "

    Limpiar_variables(ws,7,10,3,4,thin_border)
    Limpiar_variables(ws,12,13,2,5,thin_border)
    Limpiar_variables(ws,17,18,2,10,thin_border)
    Limpiar_variables(ws,24,64,2,8)
    Limpiar_variables(ws,23,24,2,7,thin_border)
    for i in range(7, 8):
        for j in range(24,64):
            copyStyle(ws.cell(row = j, column = i), ws.cell(row = 1, column = 1))


def copyStyle(newCell, cell):
    if cell.has_style:
        newCell.style = copy(cell.style)
        newCell.font = copy(cell.font)
        newCell.border = copy(cell.border)
        newCell.fill = copy(cell.fill)
        newCell.number_format = copy(cell.number_format)
        newCell.protection = copy(cell.protection)
        newCell.alignment = copy(cell.alignment)

def copyStyle_formula(ws, newCell, cell):
#     if cell.has_style:
        newCell.style = copy(cell.style)
        newCell.font = copy(cell.font)
        newCell.border = copy(cell.border)
        newCell.fill = copy(cell.fill)
        newCell.number_format = copy(cell.number_format)
        newCell.protection = copy(cell.protection)
        newCell.alignment = copy(cell.alignment)
#         print(str(cell.value), str(cell.coordinate),str(newCell.coordinate))
        ws[str(newCell.coordinate)] =openpyxl.formula.translate.Translator(cell.value,origin=str(cell.coordinate)).translate_formula(str(newCell.coordinate))
#         newCell.value = copy(cell.coordinate)
    
def style_rango_formula(ws, fila_ini, fila_fin, col_ini, col_fin):
    for i in range(col_ini, col_fin+1):
        for j in range(fila_ini+1,fila_fin):
            copyStyle_formula(ws, ws.cell(row = j, column = i), ws.cell(row = fila_ini, column = i))
#             print("Celda",j,i)
    
def style_rango(ws, fila_ini, fila_fin, col_ini, col_fin):
    for i in range(col_ini, col_fin):
        for j in range(fila_ini+1,fila_fin):
            copyStyle(ws.cell(row = j, column = i), ws.cell(row = fila_ini, column = i))
    
