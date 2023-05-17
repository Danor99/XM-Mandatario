# -*- coding: utf-8 -*-
"""
Created on Sat Apr 25 12:14:22 2020

@author: 50414
"""

from openpyxl.styles.borders import Border, Side
#from copy import copy
import pandas as pd
#import pyodbc 
from openpyxl import load_workbook
import Utilidades as uti
import os
import xlwings as xl
import numpy as np
import sys

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Datos desde el programa que genera el archivo


def verificacion_CPROG(Mes,tipo_user, mapeo_meses, Df_Agentes, directorio, directorio_archivos):

    mapeo_agente = uti.Consulta_mapeo_nombre_agente(uti.Str_desde_array(Df_Agentes['UngID'].unique()))
    Df_Agentes = pd.merge(Df_Agentes, mapeo_agente, how = 'inner', left_on ='UngID', right_on = 'objID')
    del Df_Agentes['objID']
    
    Mes_pub = mapeo_meses[Mes[5:7]]
    
    Archivo_verif =  'Cargos CPROG_'+Mes+"_"+tipo_user+"_V6.xlsx"
#    Archivo_verif_sin_form =  'Cargos CPROG_'+Mes+"_"+tipo_user+"_V5_verif.xlsx"
    Archivo_verif_sin_form = Archivo_verif
    
#    wb = load_workbook(directorio+Archivo_verif)
#    wb.save(directorio+Archivo_verif_sin_form)
        
    app = xl.App(visible=False)
    book = app.books.open(directorio+Archivo_verif_sin_form)
    book.save()
    app.kill()
        
    wb = load_workbook(directorio+Archivo_verif_sin_form)
    
    Age = 'EPID'
    Agentes = Df_Agentes['IDENTAGENT'].unique()
    
    for Age in Agentes:
    #Insumos para acceder a archivos de publicación y verificación
        
        Nombre = Df_Agentes[Df_Agentes['IDENTAGENT']==Age]['nombreCorto'].unique()[0]
        Nombre_hoja = 'CargoCPROG-'+ Age
        Archivo = Age + "-CargoCPROG-" + Mes[0:4] + Mes[5:7] + ".xlsx"
        
        #Cargar DF_ de datos en archivo de publicación
        try:
            df_AIMCP = pd.read_excel(directorio_archivos+Archivo, sheet_name = 'CargoCPROG', nrows= 24, usecols = 'B:G', skiprows= range(1,24), header = None)
            df_AIMCP.columns = ['MES','IPPM2','VenSTN','VenNI','VenI','CPROGAIM']
            df_AIMCP.dropna(axis = 0, inplace = True)
            
            df_insumos_CPROG = pd.read_excel(directorio_archivos+Archivo, sheet_name = 'CargoCPROG', nrows= 2, usecols = 'B:H', skiprows= range(1,17), header = None)
            df_insumos_CPROG.columns = ['A_NombreCorto','A_Nombre','VenSTN','VenNI','VenI', 'NMA' ,'AIMCP']
            df_insumos_CPROG.dropna(axis = 0, inplace = True)
            
            df_CPROG = pd.read_excel(directorio_archivos+Archivo, sheet_name = 'CargoCPROG', nrows= 2, usecols = 'B:E', skiprows= range(1,12), header = None)
            df_CPROG.columns = ['A_NombreCorto','CAP','Ing_Dev','CPROG']
            df_CPROG.dropna(axis = 0, inplace = True)
            
            df_IPPs = pd.read_excel(directorio_archivos+Archivo, sheet_name = 'CargoCPROG', nrows= 4, usecols = 'B:C', skiprows= range(1,6), header = None)
            df_IPPs.columns = ['Indicador','Valor']
            df_IPPs.dropna(axis = 0, inplace = True)
            
            #Cargar DF_ de datos en archivo de verificacion
            
            df_AIMCP_verif = pd.read_excel(directorio+Archivo_verif_sin_form, sheet_name = Nombre_hoja, nrows= 24, usecols = 'B:G', skiprows= range(1,22), header = None)
            df_AIMCP_verif.columns = ['MES','IPPM2','VenSTN','VenNI','VenI','CPROGAIM']
            df_AIMCP_verif.replace(' ', np.nan,inplace = True)
            df_AIMCP_verif.dropna(subset=['CPROGAIM'], axis = 0, inplace = True)
            df_AIMCP_verif.fillna(0,inplace=True)
            
            df_insumos_CPROG_verif = pd.read_excel(directorio+Archivo_verif_sin_form, sheet_name = Nombre_hoja, nrows= 2, usecols = 'B:K', skiprows= range(1,16), header = None)
            df_insumos_CPROG_verif.columns = ['A_NombreCorto','A_Nombre','VenSTN','VenNI','VenI', 'NMA' ,'AIMCP_2','CPROG','NMA_2','AIMCP']
            df_insumos_CPROG_verif.replace(' ', np.nan,inplace = True)
            df_insumos_CPROG_verif.dropna(subset=['AIMCP_2'], axis = 0, inplace = True)
            del df_insumos_CPROG_verif['NMA_2']
            del df_insumos_CPROG_verif['AIMCP_2']
            del df_insumos_CPROG_verif['CPROG']
            df_insumos_CPROG_verif.fillna(0,inplace=True)
            
            df_CPROG_verif = pd.read_excel(directorio+Archivo_verif_sin_form, sheet_name = Nombre_hoja, nrows= 2, usecols = 'B:F', skiprows= range(1,11), header = None)
            df_CPROG_verif.columns = ['A_NombreCorto','CAP','Ing_Dev','INVNUC','CPROG']
            df_CPROG_verif.replace(' ', np.nan,inplace = True)
            df_CPROG_verif.dropna(subset=['CPROG'], axis = 0, inplace = True)
            df_CPROG_verif.fillna(0,inplace=True)
            
            df_IPPs_verif = pd.read_excel(directorio+Archivo_verif_sin_form, sheet_name = Nombre_hoja, nrows= 4, usecols = 'B:C', skiprows= range(1,6), header = None)
            df_IPPs_verif.columns = ['Indicador','Valor']
            df_IPPs_verif.replace(' ', np.nan,inplace = True)
            df_IPPs_verif.dropna(subset=['Valor'], axis = 0, inplace = True)
            df_IPPs_verif.fillna(0,inplace=True)
            df_IPPs_verif.sort_values(by='Indicador', inplace=True, ascending =False)
            df_IPPs_verif.reset_index(inplace= True)
            df_IPPs.sort_values(by='Indicador',inplace=True, ascending =False)
            df_IPPs.reset_index(inplace = True)
            
            #Comparacion de datos
            #df_AIMCP_verif = df_AIMCP_verif.round({'CPROGAIM':10}).round({'CPROGAIM':9})
            #df_AIMCP_verif.CPROGAIM = df_AIMCP_verif.CPROGAIM.apply(Decimal)
            #df_AIMCP_verif.CPROGAIM = df_AIMCP_verif.CPROGAIM.apply((lambda x: x.quantize(Decimal('0.0000000001'),ROUND_HALF_UP)))
            #df_AIMCP_verif.CPROGAIM = df_AIMCP_verif.CPROGAIM.apply((lambda x: x.quantize(Decimal('0.000000001'),ROUND_HALF_UP)))
            #df_AIMCP_verif.CPROGAIM = df_AIMCP_verif.CPROGAIM.astype(float)
            if df_AIMCP.empty:
                print('No hay AIM para ', Age)
            else:
                df_AIMCP.MES = df_AIMCP.MES.dt.strftime('%Y-%m-%d')
                Equal_df_AIMCP = df_AIMCP[['IPPM2','VenSTN','VenNI','VenI']].equals(df_AIMCP_verif[['IPPM2','VenSTN','VenNI','VenI']])
                Verif_df_AIMCP = df_AIMCP_verif['CPROGAIM']-df_AIMCP['CPROGAIM']
            
            Equal_df_insumos_CPROG = df_insumos_CPROG[['A_NombreCorto', 'A_Nombre', 'VenSTN', 'VenNI', 'VenI', 'NMA']].equals(df_insumos_CPROG_verif[['A_NombreCorto', 'A_Nombre', 'VenSTN', 'VenNI', 'VenI', 'NMA']])
            Verif_df_insumos_CPROG = df_insumos_CPROG_verif['AIMCP']-df_insumos_CPROG['AIMCP']
            Equal_df_CPROG_verif = df_CPROG_verif[['A_NombreCorto','CAP','Ing_Dev']].equals(df_CPROG[['A_NombreCorto','CAP','Ing_Dev']])
            Verif_df_CPROG_verif = df_CPROG_verif['CPROG']-df_CPROG['CPROG']
            Verif_df_IPPs_verif = df_IPPs_verif['Valor']-df_IPPs['Valor']
            
            
            #Pegar datos en archivo de verificación
            ws = wb[Nombre_hoja]
            
            if df_AIMCP.empty != 1:
                uti.Copiar_DF_excel(Verif_df_AIMCP.to_frame(), ws, 23, 8, thin_border)
                
            uti.Copiar_DF_excel(Verif_df_insumos_CPROG.to_frame(), ws, 17, 12, thin_border)
            uti.Copiar_DF_excel(Verif_df_CPROG_verif.to_frame(), ws, 12, 7, thin_border)
            uti.Copiar_DF_excel(Verif_df_IPPs_verif.to_frame(), ws, 7, 4, thin_border)
            
            #Hacer log
            file = open(directorio + "Errores_CPROG.txt", "a")
            if df_AIMCP.empty != 1:
                if Equal_df_AIMCP != 1:
                        file.write( Age + " Error Insumos CPROG AIMCP" + os.linesep)
                if ((Verif_df_AIMCP.sum() >= 0.00000001) | (Verif_df_AIMCP.sum() <= -0.00000001)):
                        file.write(Age + " Error CPROG AIMCP" + os.linesep)
            if Equal_df_insumos_CPROG != 1:
                    file.write(Age + " Error Insumos CPROG" + os.linesep)
            if ((Verif_df_insumos_CPROG.sum() >= 0.00000001) | (Verif_df_insumos_CPROG.sum() <= -0.00000001)):
                    file.write(Age + " Error AIMCP" + os.linesep)
            if Equal_df_CPROG_verif != 1:
                    file.write(Age + " Error Insumos CAP IDev" + os.linesep)
            if ((Verif_df_CPROG_verif.sum() >= 0.00000001) | (Verif_df_CPROG_verif.sum() <= -0.00000001)):
                    file.write(Age + " Error CPROG" + os.linesep)
            if ((Verif_df_IPPs_verif.sum() >= 0.00000001) | (Verif_df_IPPs_verif.sum() <= -0.00000001)):
                    file.write(Age + "      Error IPPs" + os.linesep)
            file.close()
        except:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            print('\n', exc_type,'\n', exc_value,'\n', exc_traceback, '\n','No se encontro archivo  ', Archivo)
    
    wb.save(directorio+Archivo_verif_sin_form)
    

